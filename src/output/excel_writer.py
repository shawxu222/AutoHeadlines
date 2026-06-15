from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config_loader import DATA_ROOT
from src.utils.dates import compact_date


CANDIDATE_COLUMNS = [
    "candidate_id",
    "score",
    "recommended_reason",
    "title_original",
    "title_translated_candidate",
    "source",
    "source_section",
    "country_region",
    "language",
    "published_date",
    "url",
    "source_domain",
    "matched_keywords",
    "reference_similarity_score",
    "suggested_type",
    "suggested_soft_hard",
    "raw_text_preview",
    "raw_text",
    "selected",
    "notes",
    "duplicate_group",
    "extraction_warning",
]


def candidates_path(run_date: str) -> Path:
    return DATA_ROOT / "output" / f"candidates_{compact_date(run_date)}.xlsx"


def write_candidates_excel(
    candidates: list[dict[str, Any]], run_date: str, path: Path | None = None
) -> Path:
    output_path = path or candidates_path(run_date)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = []
    for candidate in candidates:
        rows.append({column: candidate.get(column, "") for column in CANDIDATE_COLUMNS})
    frame = pd.DataFrame(rows, columns=CANDIDATE_COLUMNS)
    frame.to_excel(output_path, index=False)

    _format_candidate_sheet(output_path)
    return output_path


def read_reviewed_candidates(path: Path) -> list[dict[str, Any]]:
    frame = pd.read_excel(path).fillna("")
    rows = []
    for _, row in frame.iterrows():
        item = {str(column): row[column] for column in frame.columns}
        if is_selected(item.get("selected", "")):
            rows.append(item)
    return rows


def is_selected(value: Any) -> bool:
    text = str(value).strip().lower()
    return text in {"1", "yes", "true", "是", "采纳", "y", "√", "v", "selected"}


def _format_candidate_sheet(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    sheet = workbook.active
    sheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, name="Microsoft YaHei")

    widths = {
        "A": 16,
        "B": 10,
        "C": 30,
        "D": 45,
        "E": 30,
        "F": 20,
        "G": 20,
        "H": 16,
        "I": 12,
        "J": 15,
        "K": 45,
        "L": 24,
        "M": 35,
        "N": 18,
        "O": 14,
        "P": 16,
        "Q": 60,
        "R": 12,
        "S": 24,
        "T": 16,
        "U": 28,
    }
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[letter].width = widths.get(letter, 18)
        for cell in column_cells[1:]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Microsoft YaHei", size=10)

    workbook.save(path)
